import os
import sqlite3
import threading
import time
import asyncio
import re
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
import numpy as np
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from dotenv import load_dotenv
from telegram import Bot, Update, ReplyKeyboardMarkup
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    filters,
    ContextTypes,
)
from apscheduler.schedulers.background import BackgroundScheduler
from telegram import InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import ConversationHandler


# Настройки
ADMIN_IDS = [int(id_str.strip()) for id_str in os.getenv("ADMIN_IDS", "").split(",") if id_str.strip()]

def is_admin(user_id):
    return user_id in ADMIN_IDS

load_dotenv()
TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
if not TOKEN:
    raise ValueError("Токен бота не найден в .env файле!")
# DB_NAME = "/app/data/egg_database.db"  # Для Docker
DB_NAME = "egg_database.db"  # Для локального использования


# Константа для состояния рассылки
BROADCAST_MESSAGE = 1

# Инициализация базы данных
def init_db():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    # Таблица для записей о яйценоскости
    c.execute('''CREATE TABLE IF NOT EXISTS eggs
                (id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                date TEXT,
                count INTEGER,
                notes TEXT)''')

    # Таблица для настроек пользователей
    c.execute('''CREATE TABLE IF NOT EXISTS user_settings
                (user_id INTEGER PRIMARY KEY,
                reminders_enabled BOOLEAN DEFAULT 0,
                reminder_time TEXT DEFAULT '20:00',
                timezone TEXT DEFAULT '+03:00')''')

    conn.commit()
    conn.close()

def is_valid_date(date_str):
    try:
        datetime.strptime(date_str, "%Y-%m-%d")
        return True
    except ValueError:
        return False

# Добавление записи
def add_egg_record(user_id, date, count, notes=""):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("INSERT INTO eggs (user_id, date, count, notes) VALUES (?, ?, ?, ?)",
              (user_id, date, count, notes))
    record_id = c.lastrowid
    conn.commit()
    conn.close()
    return record_id

def get_record_by_id(record_id):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT * FROM eggs WHERE id=?", (record_id,))
    record = c.fetchone()
    conn.close()
    return record

def update_record(record_id, count=None, date=None, notes=None):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    updates = []
    params = []

    if count is not None:
        updates.append("count = ?")
        params.append(count)
    if date is not None:
        updates.append("date = ?")
        params.append(date)
    if notes is not None:
        updates.append("notes = ?")
        params.append(notes)

    if updates:
        query = f"UPDATE eggs SET {', '.join(updates)} WHERE id = ?"
        params.append(record_id)
        c.execute(query, params)
        conn.commit()
    conn.close()

async def edit_entry(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        user_id = update.message.from_user.id
        args = context.args

        if len(args) < 2:
            await update.message.reply_text(
                "Используйте: /edit <ID записи> <новое количество> [дата] [комментарий]\n"
                "Пример: /edit 1 15 2023-12-20 Новый корм"
            )
            return

        record_id = int(args[0])
        new_count = int(args[1])
        new_date = args[2] if len(args) > 2 and is_valid_date(args[2]) else None
        new_notes = " ".join(args[3:]) if len(args) > 3 else None

        # Проверка, что запись принадлежит пользователю
        record = get_record_by_id(record_id)
        if not record or record[1] != user_id:
            await update.message.reply_text("❌ Запись не найдена или недоступна.")
            return

        update_record(record_id, new_count, new_date, new_notes)
        await update.message.reply_text("✅ Запись успешно обновлена!")

    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка: {str(e)}")

def delete_record(record_id):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("DELETE FROM eggs WHERE id=?", (record_id,))
    conn.commit()
    conn.close()

async def delete_entry(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        user_id = update.message.from_user.id
        args = context.args

        if not args:
            await update.message.reply_text("Используйте: /delete <ID записи>")
            return

        record_id = int(args[0])

        # Проверка, что запись принадлежит пользователю
        record = get_record_by_id(record_id)
        if not record or record[1] != user_id:
            await update.message.reply_text("❌ Запись не найдена или недоступна.")
            return

        delete_record(record_id)
        await update.message.reply_text("✅ Запись успешно удалена!")

    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка: {str(e)}")

# Получение статистики
def get_stats(user_id, days=7):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    start_date = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")

    # Запрос для получения данных
    query = '''SELECT id, date, count
               FROM eggs
               WHERE user_id = ? AND date >= ?
               ORDER BY date'''
    c.execute(query, (user_id, start_date))
    data = c.fetchall()
    conn.close()

    # Группируем данные по дате и суммируем количество яиц
    stats = {}
    for record_id, date, count in data:
        if date in stats:
            stats[date]['total'] += count
            stats[date]['ids'].append(record_id)
        else:
            stats[date] = {'total': count, 'ids': [record_id]}

    # Преобразуем в список для удобства
    result = [(date, details['total'], details['ids']) for date, details in stats.items()]
    return result

async def show_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    days = int(context.args[0]) if context.args else 7
    data = get_stats(user_id, days)

    if not data:
        await update.message.reply_text("❌ Нет данных за указанный период.")
        return

    stats_text = f"📊 Ваша статистика за {days} дней:\n"
    total = 0

    for date, count, ids in data:
        stats_text += f"📅 {date}: {count} яиц\n"
        stats_text += f"   ID записей: {', '.join(map(str, ids))}\n"
        total += count

    stats_text += f"\nВсего: {total} яиц\nСреднее: {total/len(data):.1f} яиц/день"
    await update.message.reply_text(stats_text)

def has_today_entry(user_id):
    today = datetime.now().strftime("%Y-%m-%d")
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM eggs WHERE user_id = ? AND date = ?",
              (user_id, today))
    count = c.fetchone()[0]
    conn.close()
    return count > 0

# Функция для генерации графиков
def generate_plot(user_id, days=7):
    data = get_stats(user_id, days)
    if not data:
        return None

    # Извлекаем даты и значения
    dates = [datetime.strptime(row[0], "%Y-%m-%d") for row in data]
    counts = [row[1] for row in data]

    plt.figure(figsize=(10, 6))
    plt.plot(dates, counts, marker='o', linestyle='-', color='#ff6b6b')
    plt.title(f'Ваша яйценоскость за {days} дней')
    plt.xlabel('Дата')
    plt.ylabel('Количество яиц')
    plt.grid(True, alpha=0.3)
    plt.xticks(rotation=45)
    plt.tight_layout()

    filename = f"egg_stats_{user_id}_{days}days.png"
    plt.savefig(filename, dpi=100)
    plt.close()
    return filename

# Команда для графиков
async def show_graph(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        user_id = update.message.from_user.id
        days = int(context.args[0]) if context.args else 7
        if days <= 0:
            raise ValueError

        filename = generate_plot(user_id, days)
        if filename:
            await update.message.reply_photo(
                photo=open(filename, 'rb'),
                caption=f'📈 График яйценоскости за {days} дней'
            )
            os.remove(filename)
        else:
            await update.message.reply_text("❌ Нет данных для построения графика")

    except (ValueError, IndexError):
        await update.message.reply_text("Используйте: /graph <количество_дней> (по умолчанию 7)")
    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка: {str(e)}")

# Функция аналитики
def calculate_analytics(user_id, days=7):
    # Получаем данные за два периода для сравнения
    data = get_stats(user_id, days * 2)

    if len(data) < 2:
        return None

    # Разделяем данные на текущий и предыдущий периоды
    current = data[-days:]
    previous = data[:-days]

    # Основные метрики
    current_counts = [c[1] for c in current]
    avg_current = np.mean(current_counts)
    avg_previous = np.mean([p[1] for p in previous]) if previous else 0

    # Статистический анализ (тренд)
    slope, _, _, _, _ = stats.linregress(
        range(len(current_counts)), current_counts
    )

    # Анализ заметок
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute('''SELECT notes FROM eggs
                 WHERE user_id = ? AND date >= date('now', ?)''',
              (user_id, f'-{days} days',))
    notes = [note[0].lower() for note in c.fetchall() if note[0]]

    word_analysis = {}
    for note in notes:
        for word in note.split():
            word_analysis[word] = word_analysis.get(word, 0) + 1
    top_words = sorted(word_analysis.items(),
                      key=lambda x: x[1], reverse=True)[:3]

    conn.close()

    return {
        'current_avg': avg_current,
        'previous_avg': avg_previous,
        'trend': slope * days,  # Общий тренд за период
        'max_day': max(current, key=lambda x: x[1]),
        'min_day': min(current, key=lambda x: x[1]),
        'top_words': top_words
    }

async def show_analytics(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        user_id = update.message.from_user.id
        days = int(context.args[0]) if context.args else 7
        analytics = calculate_analytics(user_id, days)

        if not analytics:
            await update.message.reply_text("❌ Недостаточно данных для анализа")
            return

        response = (
            f"📈 Ваша аналитика за {days} дней:\n\n"
            f"▪ Среднее: {analytics['current_avg']:.1f} яиц/день\n"
            f"▪ Тренд: {'↑' if analytics['trend'] > 0 else '↓'} "
            f"{abs(analytics['trend']):.1f} яиц за период\n"
            f"▪ Рекорд: {analytics['max_day'][1]} яиц ({analytics['max_day'][0]})\n"
            f"▪ Минимум: {analytics['min_day'][1]} яиц ({analytics['min_day'][0]})\n"
        )

        if analytics['previous_avg']:
            change = ((analytics['current_avg'] - analytics['previous_avg']) /
                     analytics['previous_avg'] * 100)
            response += (
                f"\n🔄 Изменение к прошлому периоду: "
                f"{change:+.1f}%\n"
            )

        if analytics['top_words']:
            response += "\n🔍 Частые упоминания:\n" + "\n".join(
                [f"- {word} ({count} раз)" for word, count in analytics['top_words']]
            )

        await update.message.reply_text(response)

    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка: {str(e)}")

# Команды бота
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        ["/add", "/edit", "/delete"],
        ["/stats", "/graph", "/analytics"],
        ["/export", "/myid"],
        ["/help", "/reminders", "/donate ☕"],
    ]

    commands_text = (
        "🐔 Бот для учета яйценоскости кур!\n\n"
        "Основные команды:\n"
        "▪ /add — добавить запись\n"
        "▪ /stats [дни] — статистика\n"
        "▪ /graph [дни] — график\n"
        "▪ /analytics [дни] — расширенная аналитика\n\n"
        "Управление записями:\n"
        "▪ /edit <ID> <количество> [дата] [комментарий] — изменить запись\n"
        "▪ /delete <ID> — удалить запись\n\n"
        "Экспорт данных:\n"
        "▪ /export [дни] — выгрузить данные в Excel\n\n"
        "▪ /help — список всех команд с кратким описанием\n\n"
        "Управление напоминаниями:\n"
        "▪ /reminders — управлять напоминаниями 🔔\n\n"
        "Поддержать проект:\n"
        "▪ /donate — оплатить чашку кофе ☕\n\n"
        "Используйте кнопки ниже или введите команду вручную."
    )

    await update.message.reply_text(
        commands_text,
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    )

async def add_entry(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Введите данные в формате:\n"
        "<количество яиц> [дата] [комментарий]\n\n"
        "Примеры:\n"
        "12 — добавить 12 яиц на сегодня\n"
        "12 2023-12-15 — добавить 12 яиц на 15 декабря 2023\n"
        "12 сегодня Корм поменяли — добавить 12 яиц на сегодня с комментарием"
    )

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    args = context.args
    if not args:
        help_text = (
            "🐔 Бот для учёта яйценоскости кур!\n\n"
            "Список команд:\n"
            "▪ /add — добавить запись\n"
            "▪ /stats — статистика\n"
            "▪ /graph — график\n"
            "▪ /analytics — аналитика\n"
            "▪ /edit — изменить запись\n"
            "▪ /delete — удалить запись\n"
            "▪ /export — экспорт\n"
            "▪ /help — справка\n"
            "▪ /donate — поддержка проекта\n"
            "▪ /myid — показать ваш Telegram ID\n"
            "▪ /reminders — управление напоминаниями\n\n"
            
            "Подробности по каждой команде смотрите через '/help <название команды>'"
        )
    else:
        # Подробное описание конкретной команды
        command = args[0].lower()
        if command == "add":
            help_text = (
                "📝 Добавление записи:\n"
                "Используйте команду /add или введите данные в формате:\n"
                "<количество яиц> [дата] [комментарий]\n\n"
                "Примеры:\n"
                "12 — добавить 12 яиц на сегодня\n"
                "12 2023-12-15 — добавить 12 яиц на 15 декабря 2023\n"
                "12 сегодня Корм поменяли — добавить 12 яиц на сегодня с комментарием"
            )
        elif command == "stats":
            help_text = (
                "📊 Статистика:\n"
                "Используйте команду /stats [дни], чтобы увидеть статистику за указанное количество дней.\n\n"
                "Примеры:\n"
                "/stats — статистика за 7 дней\n"
                "/stats 14 — статистика за 14 дней"
            )
        elif command == "graph":
            help_text = (
                "📈 График:\n"
                "Используйте команду /graph [дни], чтобы построить график яйценоскости.\n\n"
                "Примеры:\n"
                "/graph — график за 7 дней\n"
                "/graph 30 — график за 30 дней"
            )
        elif command == "analytics":
            help_text = (
                "📈 Аналитика:\n"
                "Используйте команду /analytics [дни], чтобы получить аналитику за указанное количество дней.\n\n"
                "Примеры:\n"
                "/analytics — аналитика за 7 дней\n"
                "/analytics 14 — аналитика за 14 дней"
            )
        elif command == "edit":
            help_text = (
                "✏️ Редактирование записи:\n"
                "Используйте команду /edit <ID> <количество> [дата] [комментарий], чтобы изменить запись.\n\n"
                "Примеры:\n"
                "/edit 1 15 — изменить количество яиц на 15 для записи с ID 1\n"
                "/edit 1 15 2023-12-20 Новый корм — изменить запись с ID 1"
            )
        elif command == "delete":
            help_text = (
                "❌ Удаление записи:\n"
                "Используйте команду /delete <ID>, чтобы удалить запись.\n\n"
                "Примеры:\n"
                "/delete 1 — удалить запись с ID 1"
            )
        elif command == "export":
            help_text = (
                "⬇️ Экспорт в Excel:\n"
                "Используйте команду /export [дни], чтобы получить файл со статистикой.\n\n"
                "Примеры:\n"
                "/export 5 — получить файл Excel со статистикой за 5 дней\n"
                "/export 2025-01-23 2025-02-06 — получить файл Excel со статистикой за указанный период\n"
            )
        elif command == "reminders":
            help_text = (
                "🔔 Управление напоминаниями:\n"
                "Используйте команду /reminders для настройки ежедневных напоминаний о внесении данных о яйценоскости.\n\n"
                "Примеры:\n"
                "/reminders on — включить напоминания\n"
                "/reminders off — отключить напоминания\n"
                "/reminders time ЧЧ:ММ — задать время напоминания (например, /reminders time 19:00)\n"
                "/reminders — посмотреть текущее состояние напоминаний"
            )
        else:
            help_text = f"❌ Команда '{command}' не найдена. Используйте /help для списка команд."

    await update.message.reply_text(help_text)

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):

    # Если ожидается сообщение для рассылки - пропускаем обычную обработку
    if context.user_data.get('awaiting_broadcast'):
        return
    
    try:
        user_id = update.message.from_user.id
        text = update.message.text

         # Пропускаем административные команды
        admin_commands = ["📊 Общая статистика", "👥 Список пользователей", "📢 Рассылка"]
        if text in admin_commands:
            return
        
        if text.startswith('/'):
            return

        parts = text.split(maxsplit=2)
        count = int(parts[0])  # Количество яиц

        # Обработка даты
        if len(parts) > 1 and parts[1].lower() != "сегодня":
            if not is_valid_date(parts[1]):
                await update.message.reply_text("❌ Неверный формат даты! Используйте ГГГГ-ММ-ДД.")
                return
            date = parts[1]
            notes = parts[2] if len(parts) > 2 else ""
        else:
            date = datetime.now().strftime("%Y-%m-%d")
            notes = parts[2] if len(parts) > 2 else ""

        # Добавление записи
        record_id = add_egg_record(user_id, date, count, notes)
        await update.message.reply_text(
            f"✅ Добавлено: {count} яиц\n"
            f"Дата: {date}\n"
            f"Заметка: {notes}\n"
            f"ID записи: {record_id}"
        )

    except Exception as e:
        await update.message.reply_text(
            "❌ Ошибка формата! Примеры:\n"
            "12 — добавить 12 яиц на сегодня\n"
            "12 2023-12-15 — добавить 12 яиц на 15 декабря 2023\n"
            "12 сегодня Корм поменяли — добавить 12 яиц на сегодня с комментарием"
        )

# Выгрузка в Excel
def export_to_excel(user_id, start_date=None, end_date=None):
    # Получаем данные из базы
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    query = '''SELECT date, SUM(count), GROUP_CONCAT(id)
               FROM eggs
               WHERE user_id = ?
               AND date >= ?
               AND date <= ?
               GROUP BY date
               ORDER BY date'''
    start_date = start_date or (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
    end_date = end_date or datetime.now().strftime("%Y-%m-%d")

    c.execute(query, (user_id, start_date, end_date))
    data = c.fetchall()
    conn.close()

    if not data:
        return None

    # Создаем Excel-файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Яйценоскость"
    ws.append(["Дата", "Количество яиц", "ID записей"])

    # Применяем стили к заголовкам
    for col in range(1, 4):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

   # Заполняем данные
    for date, count, ids in data:
        ws.append([date, count, ", ".join(map(str, ids))])

    # Выравниваем ID записей по правому краю
    right_alignment = Alignment(horizontal="right")
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = right_alignment

    # Автоширина столбцов
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

    # Добавляем график
    chart = LineChart()
    chart.title = "Яйценоскость"
    chart.x_axis.title = "Дата"
    chart.y_axis.title = "Количество яиц"

    data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(data) + 1)
    categories_ref = Reference(ws, min_col=1, min_row=2, max_row=len(data) + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(categories_ref)

    ws.add_chart(chart, "E2")

    # Сохраняем файл
    filename = f"egg_stats_{user_id}_{start_date}_to_{end_date}.xlsx"
    wb.save(filename)
    return filename

async def export_data(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        user_id = update.message.from_user.id
        args = context.args

        if len(args) == 2:  # Если указаны даты начала и конца
            start_date, end_date = args
            if not (is_valid_date(start_date) and is_valid_date(end_date)):
                await update.message.reply_text("❌ Неверный формат даты! Используйте ГГГГ-ММ-ДД.")
                return
        else:
            start_date = None
            end_date = None

        # Выгружаем данные в Excel
        filename = export_to_excel(user_id, start_date, end_date)
        if filename:
            await update.message.reply_document(
                document=open(filename, 'rb'),
                caption="Можете скачать файл с таблицей"
            )
            os.remove(filename)  # Удаляем временный файл
        else:
            await update.message.reply_text("❌ Нет данных для выгрузки.")

    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка: {str(e)}")

async def donate(update: Update, context: ContextTypes.DEFAULT_TYPE):
    donate_url = "https://pay.cloudtips.ru/p/dbed3f9a"  # Замените на вашу ссылку
    message = (
        "☕ Если вам нравится этот бот, вы можете поддержать его разработку!\n\n"
        f"👉 [Оплатить чашку кофе]({donate_url})"
    )
    await update.message.reply_text(
        message,
        parse_mode="Markdown",
        disable_web_page_preview=True
    )

async def show_my_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    await update.message.reply_text(f"🆔 Ваш Telegram ID: `{user_id}`", parse_mode="Markdown")

async def send_reminder_async(bot, user_id):
    try:
        await bot.send_message(
            chat_id=user_id,
            text="⏰ Напоминание! Сегодня вы еще не вносили данные о яйцах.\n"
                 "Используйте команду /add или просто отправьте число."
        )
    except Exception as e:
        print(f"Ошибка при отправке напоминания пользователю {user_id}: {str(e)}")

def send_reminder(user_id):
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    loop.run_until_complete(send_reminder_async(Bot(token=TOKEN), user_id))
    loop.close()

# -------------
def check_and_remind():
    """Синхронная функция для проверки и отправки напоминаний с учетом часового пояса"""
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT user_id, reminder_time, timezone FROM user_settings WHERE reminders_enabled=1")
    users = c.fetchall()
    
    for user_id, reminder_time, timezone in users:
        try:
            if has_today_entry(user_id):
                continue
                
            hour, minute = map(int, reminder_time.split(':'))
            
            # Получаем текущее время с учетом часового пояса пользователя
            try:
                tz_offset = int(timezone[:3])  # Преобразуем "+03:00" в 3
            except:
                tz_offset = 3  # Значение по умолчанию, если часовой пояс некорректен
                
            now = datetime.utcnow() + timedelta(hours=tz_offset)
            
            if now.hour == hour and now.minute == minute:
                threading.Thread(target=send_reminder, args=(user_id,)).start()
        except Exception as e:
            print(f"Ошибка при проверке напоминания для {user_id}: {str(e)}")
    
    conn.close()

def start_scheduler():
    """Запуск планировщика в фоновом режиме"""
    scheduler = BackgroundScheduler()
    scheduler.add_job(check_and_remind, 'interval', minutes=1)
    scheduler.start()
    print("Планировщик напоминаний запущен")
    
    # Бесконечный цикл, чтобы поток не завершался
    try:
        while True:
            time.sleep(60)
    except KeyboardInterrupt:
        scheduler.shutdown()

# Функции для управления напоминаниями
def get_user_settings(user_id):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT reminders_enabled, reminder_time, timezone FROM user_settings WHERE user_id=?", (user_id,))
    settings = c.fetchone()
    conn.close()
    return settings or (False, '20:00', '+03:00')  # Возвращаем время и часовой пояс по умолчанию

def update_user_settings(user_id, reminders_enabled=None, reminder_time=None, timezone=None):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    
    if not c.execute("SELECT 1 FROM user_settings WHERE user_id=?", (user_id,)).fetchone():
        c.execute("INSERT INTO user_settings (user_id) VALUES (?)", (user_id,))
    
    updates = []
    params = []
    
    if reminders_enabled is not None:
        updates.append("reminders_enabled = ?")
        params.append(reminders_enabled)
    if reminder_time is not None:
        updates.append("reminder_time = ?")
        params.append(reminder_time)
    if timezone is not None:
        updates.append("timezone = ?")
        params.append(timezone)
    
    if updates:
        query = f"UPDATE user_settings SET {', '.join(updates)} WHERE user_id = ?"
        params.append(user_id)
        c.execute(query, params)
    
    conn.commit()
    conn.close()

async def manage_reminders(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message is None or update.message.from_user is None:
        return  # Или отправить сообщение о том, что произошла ошибка
    
    user_id = update.message.from_user.id
    args = context.args
    reminders_enabled, reminder_time, timezone = get_user_settings(user_id)

    if not args:
        status = "включены" if reminders_enabled else "выключены"
        await update.message.reply_text(
            f"🔔 Текущие настройки напоминаний:\n"
            f"Статус: {status}\n"
            f"Время: {reminder_time} (по вашему времени UTC{timezone})\n\n"
            "Доступные команды:\n"
            "/reminders - статус напоминаний\n"
            "/reminders on - включить\n"
            "/reminders off - выключить\n"
            "/reminders time ЧЧ:ММ - установить время\n"
            "/reminders tz ±ЧЧ:ММ - установить часовой пояс\n\n"
            "Примеры:\n"
            "/reminders time 19:00 - напоминание в 19:00\n"
            "/reminders tz +05:00 - часовой пояс UTC+5"
        )
        return

    action = args[0].lower()
    if action == "on":
        update_user_settings(user_id, reminders_enabled=True)
        await update.message.reply_text("🔔 Напоминания включены!")
    elif action == "off":
        update_user_settings(user_id, reminders_enabled=False)
        await update.message.reply_text("🔕 Напоминания выключены!")
    elif action == "time" and len(args) > 1:
        try:
            # Проверка формата времени
            datetime.strptime(args[1], "%H:%M")
            update_user_settings(user_id, reminder_time=args[1])
            await update.message.reply_text(
                f"⏰ Время напоминания установлено на {args[1]} (UTC{timezone})"
            )
        except ValueError:
            await update.message.reply_text("❌ Неверный формат времени! Используйте ЧЧ:ММ")
    elif action == "tz" and len(args) > 1:
        try:
            # Проверка формата часового пояса
            tz = args[1]
            if not re.match(r'^[+-]\d{2}:\d{2}$', tz):
                raise ValueError
            update_user_settings(user_id, timezone=tz)
            await update.message.reply_text(
                f"🌍 Часовой пояс установлен на UTC{tz}\n"
                f"Теперь напоминания будут приходить в {reminder_time} по вашему времени"
            )
        except ValueError:
            await update.message.reply_text("❌ Неверный формат часового пояса! Используйте ±ЧЧ:ММ (например +03:00 или -05:00)")
    else:
        await update.message.reply_text("❌ Неверная команда!")

# Функция для формирования клавиатуры
def create_reply_keyboard():
    keyboard = [
        ['/add', '/edit', '/delete'],
        ['/stats', '/graph', '/analytics'],
        ['/export', '/myid', '/help'],
        ['/reminders', '/donate ☕']
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
# ______________________________________________________________________________________________
# Получение общей статистики
def get_general_stats():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    
    c.execute("SELECT COUNT(DISTINCT user_id) FROM eggs")
    total_users = c.fetchone()[0]
    
    c.execute("SELECT COUNT(*), SUM(count) FROM eggs")
    total_records, total_eggs = c.fetchone()
    
    c.execute("SELECT COUNT(DISTINCT user_id) FROM eggs WHERE date >= date('now', '-7 days')")
    active_users = c.fetchone()[0]
    
    conn.close()
    
    return {
        "total_users": total_users or 0,
        "total_records": total_records or 0,
        "total_eggs": total_eggs or 0,
        "active_users": active_users or 0
    }

# Показать общую статистику
async def show_general_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.message.from_user.id):
        return
    
    stats = get_general_stats()
    response = (
        "📊 Общая статистика:\n\n"
        f"• Всего пользователей: {stats['total_users']}\n"
        f"• Всего записей: {stats['total_records']}\n"
        f"• Всего яиц: {stats['total_eggs']}\n"
        f"• Активных пользователей (7 дней): {stats['active_users']}"
    )
    await update.message.reply_text(response)

# Получить список пользователей
async def list_users(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.message.from_user.id):
        return
    
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT DISTINCT user_id, COUNT(*) as entries FROM eggs GROUP BY user_id ORDER BY entries DESC")
    users = c.fetchall()
    conn.close()
    
    if not users:
        await update.message.reply_text("❌ Нет данных о пользователях")
        return
    
    response = "👥 Список пользователей:\n\n"
    for idx, (user_id, entries) in enumerate(users, 1):
        response += f"{idx}. ID: {user_id} - Записей: {entries}\n"
    
    await update.message.reply_text(response)

# Рассылка сообщений
async def broadcast_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    if not is_admin(user_id):
        return
    
    # Запрашиваем сообщение для рассылки
    await update.message.reply_text("Введите сообщение для рассылки:")
    
    # Устанавливаем состояние ожидания сообщения
    return BROADCAST_MESSAGE  # Возвращаем состояние

async def handle_broadcast_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    message = update.message.text
    
    # Получаем список пользователей
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT DISTINCT user_id FROM eggs")
    user_ids = [row[0] for row in c.fetchall()]
    conn.close()
    
    success = 0
    failed = 0
    
    for uid in user_ids:
        try:
            await context.bot.send_message(
                chat_id=uid,
                text=f"📢 Сообщение от администратора:\n\n{message}"
            )
            success += 1
        except Exception:
            failed += 1
    
    await update.message.reply_text(
        f"✅ Рассылка завершена!\n"
        f"Успешно: {success}\n"
        f"Не удалось: {failed}"
    )
    
    # Сбрасываем состояние
    return ConversationHandler.END

# Обработчик рассылки
async def handle_broadcast(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    if not is_admin(user_id) or context.user_data.get('awaiting_broadcast') != True:
        return
    
    message = update.message.text
    context.user_data.pop('awaiting_broadcast', None)  # Сразу очищаем флаг
    
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT DISTINCT user_id FROM eggs")
    user_ids = [row[0] for row in c.fetchall()]
    conn.close()
    
    success = 0
    failed = 0
    
    for uid in user_ids:
        try:
            await context.bot.send_message(
                chat_id=uid,
                text=f"📢 Сообщение от администратора:\n\n{message}"
            )
            success += 1
        except Exception:
            failed += 1
    
    await update.message.reply_text(
        f"✅ Рассылка завершена!\n"
        f"Успешно: {success}\n"
        f"Не удалось: {failed}"
    )

# Панель администратора
async def admin_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    if not is_admin(user_id):
        await update.message.reply_text("❌ Доступ запрещен!")
        return
    
    keyboard = [
        ["📊 Общая статистика", "👥 Список пользователей"],
        ["📢 Рассылка"]
    ]
    await update.message.reply_text(
        "Панель администратора:",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=True)
    )

async def cancel_broadcast(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.pop('awaiting_broadcast', None)
    await update.message.reply_text("❌ Рассылка отменена")
    return ConversationHandler.END

# Основная функция
def main():
    """Основная функция для запуска бота"""
    init_db()
    scheduler_thread = threading.Thread(target=start_scheduler, daemon=True)
    scheduler_thread.start()
    
    application = Application.builder().token(TOKEN).build()

    # Добавляем ConversationHandler для рассылки
    conv_handler = ConversationHandler(
        entry_points=[
            MessageHandler(filters.Text(["📢 Рассылка"]) & filters.ChatType.PRIVATE, broadcast_message)
        ],
        states={
            BROADCAST_MESSAGE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_broadcast_message)
            ]
        },
        fallbacks=[CommandHandler("cancel", cancel_broadcast)],  # Добавить функцию отмены
        map_to_parent={
            ConversationHandler.END: ConversationHandler.END
        }
    )
    
    application.add_handler(conv_handler)

    # Добавляем обработчики администратора ВЫШЕ обычных
    application.add_handler(CommandHandler("admin", admin_panel))
    application.add_handler(MessageHandler(
        filters.Text(["📊 Общая статистика"]) & filters.ChatType.PRIVATE, 
        show_general_stats
    ))
    application.add_handler(MessageHandler(
        filters.Text(["👥 Список пользователей"]) & filters.ChatType.PRIVATE, 
        list_users
    ))

    # Добавляем обработчики пользовательских команд
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("add", add_entry))
    application.add_handler(CommandHandler("edit", edit_entry))
    application.add_handler(CommandHandler("delete", delete_entry))
    application.add_handler(CommandHandler("stats", show_stats))
    application.add_handler(CommandHandler("graph", show_graph))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    application.add_handler(CommandHandler("analytics", show_analytics))
    application.add_handler(CommandHandler("export", export_data))
    application.add_handler(CommandHandler("help", help_command))
    application.add_handler(CommandHandler("donate", donate))
    application.add_handler(CommandHandler("myid", show_my_id))
    application.add_handler(CommandHandler("reminders", manage_reminders))
    
    # Добавить функцию отмены рассылки

    # Запускаем бота в режиме опроса
    print("Бот запущен. Ожидание сообщений...")
    application.run_polling()

if __name__ == "__main__":
    main()