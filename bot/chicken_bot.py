import os
from dotenv import load_dotenv
import sqlite3
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes
import numpy as np
from scipy import stats
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

# Настройки
load_dotenv()

TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
if not TOKEN:
    raise ValueError("Токен бота не найден в .env файле!")
DB_NAME = "/data/egg_database.db"

# Инициализация базы данных
def init_db():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    # Таблица для записей о яйценоскости
    c.execute('''CREATE TABLE IF NOT EXISTS eggs
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  user_id INTEGER,
                  date TEXT,
                  count INTEGER,
                  notes TEXT)''')

    # Таблица для настроек пользователей
    c.execute('''CREATE TABLE IF NOT EXISTS user_settings
                 (user_id INTEGER PRIMARY KEY,
                  reminders_enabled BOOLEAN DEFAULT 0,
                  reminder_time TEXT DEFAULT '20:00')''')

    conn.commit()
    conn.close()

def is_valid_date(date_str):
    try:
        datetime.strptime(date_str, "%Y-%m-%d")
        return True
    except ValueError:
        return False

# Добавление записи
def add_egg_record(user_id, date, count, notes=""):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("INSERT INTO eggs (user_id, date, count, notes) VALUES (?, ?, ?, ?)",
              (user_id, date, count, notes))
    record_id = c.lastrowid  # Получаем ID новой записи
    conn.commit()
    conn.close()
    return record_id

def get_record_by_id(record_id):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT * FROM eggs WHERE id=?", (record_id,))
    record = c.fetchone()
    conn.close()
    return record

def update_record(record_id, count=None, date=None, notes=None):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    updates = []
    params = []

    if count is not None:
        updates.append("count = ?")
        params.append(count)
    if date is not None:
        updates.append("date = ?")
        params.append(date)
    if notes is not None:
        updates.append("notes = ?")
        params.append(notes)

    if updates:
        query = f"UPDATE eggs SET {', '.join(updates)} WHERE id = ?"
        params.append(record_id)
        c.execute(query, params)
        conn.commit()

    conn.close()

async def edit_entry(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        user_id = update.message.from_user.id
        args = context.args

        if len(args) < 2:
            await update.message.reply_text(
                "Используйте: /edit <ID записи> <новое количество> [дата] [комментарий]\n"
                "Пример: /edit 1 15 2023-12-20 Новый корм"
            )
            return

        record_id = int(args[0])
        new_count = int(args[1])
        new_date = args[2] if len(args) > 2 and is_valid_date(args[2]) else None
        new_notes = " ".join(args[3:]) if len(args) > 3 else None

        # Проверка, что запись принадлежит пользователю
        record = get_record_by_id(record_id)
        if not record or record[1] != user_id:
            await update.message.reply_text("❌ Запись не найдена или недоступна.")
            return

        update_record(record_id, new_count, new_date, new_notes)
        await update.message.reply_text("✅ Запись успешно обновлена!")

    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка: {str(e)}")

def delete_record(record_id):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("DELETE FROM eggs WHERE id=?", (record_id,))
    conn.commit()
    conn.close()

async def delete_entry(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        user_id = update.message.from_user.id
        args = context.args

        if not args:
            await update.message.reply_text("Используйте: /delete <ID записи>")
            return

        record_id = int(args[0])

        # Проверка, что запись принадлежит пользователю
        record = get_record_by_id(record_id)
        if not record or record[1] != user_id:
            await update.message.reply_text("❌ Запись не найдена или недоступна.")
            return

        delete_record(record_id)
        await update.message.reply_text("✅ Запись успешно удалена!")

    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка: {str(e)}")

# Получение статистики
def get_stats(user_id, days=7):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    start_date = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")

    # Запрос для получения данных
    query = '''SELECT id, date, count
               FROM eggs
               WHERE user_id = ? AND date >= ?
               ORDER BY date'''
    c.execute(query, (user_id, start_date))
    data = c.fetchall()
    conn.close()

    # Группируем данные по дате и суммируем количество яиц
    stats = {}
    for record_id, date, count in data:
        if date in stats:
            stats[date]['total'] += count
            stats[date]['ids'].append(record_id)
        else:
            stats[date] = {'total': count, 'ids': [record_id]}

    # Преобразуем в список для удобства
    result = [(date, details['total'], details['ids']) for date, details in stats.items()]
    return result

async def show_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    days = int(context.args[0]) if context.args else 7
    data = get_stats(user_id, days)

    if not data:
        await update.message.reply_text("❌ Нет данных за указанный период.")
        return

    stats_text = f"📊 Ваша статистика за {days} дней:\n"
    total = 0

    for date, count, ids in data:
        stats_text += f"📅 {date}: {count} яиц\n"
        stats_text += f"   ID записей: {', '.join(map(str, ids))}\n"
        total += count

    stats_text += f"\nВсего: {total} яиц\nСреднее: {total/len(data):.1f} яиц/день"
    await update.message.reply_text(stats_text)

def has_today_entry(user_id):
    today = datetime.now().strftime("%Y-%m-%d")
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM eggs WHERE user_id = ? AND date = ?",
              (user_id, today))
    count = c.fetchone()[0]
    conn.close()
    return count > 0


# Функция для генерации графиков
def generate_plot(user_id, days=7):
    data = get_stats(user_id, days)
    if not data:
        return None

    # Извлекаем даты и значения
    dates = [datetime.strptime(row[0], "%Y-%m-%d") for row in data]
    counts = [row[1] for row in data]

    plt.figure(figsize=(10, 6))
    plt.plot(dates, counts, marker='o', linestyle='-', color='#ff6b6b')
    plt.title(f'Ваша яйценоскость за {days} дней')
    plt.xlabel('Дата')
    plt.ylabel('Количество яиц')
    plt.grid(True, alpha=0.3)
    plt.xticks(rotation=45)
    plt.tight_layout()

    filename = f"egg_stats_{user_id}_{days}days.png"
    plt.savefig(filename, dpi=100)
    plt.close()
    return filename

# Команда для графиков
async def show_graph(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        user_id = update.message.from_user.id  # Получаем ID пользователя
        days = int(context.args[0]) if context.args else 7
        if days <= 0:
            raise ValueError

        filename = generate_plot(user_id, days)  # Передаем user_id
        if filename:
            await update.message.reply_photo(
                photo=open(filename, 'rb'),
                caption=f'📈 График яйценоскости за {days} дней'
            )
            os.remove(filename)
        else:
            await update.message.reply_text("❌ Нет данных для построения графика")

    except (ValueError, IndexError):
        await update.message.reply_text("Используйте: /graph <количество_дней> (по умолчанию 7)")
    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка: {str(e)}")

# Новая функция аналитики


# Новая команда аналитики
def calculate_analytics(user_id, days=7):
    # Получаем данные за два периода для сравнения
    data = get_stats(user_id, days * 2)

    if len(data) < 2:
        return None

    # Разделяем данные на текущий и предыдущий периоды
    current = data[-days:]
    previous = data[:-days]

    # Основные метрики
    current_counts = [c[1] for c in current]
    avg_current = np.mean(current_counts)
    avg_previous = np.mean([p[1] for p in previous]) if previous else 0

    # Статистический анализ (тренд)
    slope, _, _, _, _ = stats.linregress(
        range(len(current_counts)), current_counts
    )

    # Анализ заметок
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute('''SELECT notes FROM eggs
                 WHERE user_id = ? AND date >= date('now', ?)''',
              (user_id, f'-{days} days',))
    notes = [note[0].lower() for note in c.fetchall() if note[0]]

    word_analysis = {}
    for note in notes:
        for word in note.split():
            word_analysis[word] = word_analysis.get(word, 0) + 1
    top_words = sorted(word_analysis.items(),
                      key=lambda x: x[1], reverse=True)[:3]

    conn.close()

    return {
        'current_avg': avg_current,
        'previous_avg': avg_previous,
        'trend': slope * days,  # Общий тренд за период
        'max_day': max(current, key=lambda x: x[1]),
        'min_day': min(current, key=lambda x: x[1]),
        'top_words': top_words
    }

async def show_analytics(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        user_id = update.message.from_user.id
        days = int(context.args[0]) if context.args else 7
        analytics = calculate_analytics(user_id, days)

        if not analytics:
            await update.message.reply_text("❌ Недостаточно данных для анализа")
            return

        response = (
            f"📈 Ваша аналитика за {days} дней:\n\n"
            f"▪ Среднее: {analytics['current_avg']:.1f} яиц/день\n"
            f"▪ Тренд: {'↑' if analytics['trend'] > 0 else '↓'} "
            f"{abs(analytics['trend']):.1f} яиц за период\n"
            f"▪ Рекорд: {analytics['max_day'][1]} яиц ({analytics['max_day'][0]})\n"
            f"▪ Минимум: {analytics['min_day'][1]} яиц ({analytics['min_day'][0]})\n"
        )

        if analytics['previous_avg']:
            change = ((analytics['current_avg'] - analytics['previous_avg']) /
                     analytics['previous_avg'] * 100)
            response += (
                f"\n🔄 Изменение к прошлому периоду: "
                f"{change:+.1f}%\n"
            )

        if analytics['top_words']:
            response += "\n🔍 Частые упоминания:\n" + "\n".join(
                [f"- {word} ({count} раз)" for word, count in analytics['top_words']]
            )

        await update.message.reply_text(response)

    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка: {str(e)}")

# Команды бота
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        ["/add", "/edit", "/delete"],
        ["/stats", "/graph", "/analytics"],
        ["/export"],
        ["/help", "/donate ☕"],

    ]

    commands_text = (
        "🐔 Бот для учета яйценоскости кур!\n\n"
        "Основные команды:\n"
        "▪ /add — добавить запись\n"
        "▪ /stats [дни] — статистика\n"
        "▪ /graph [дни] — график\n"
        "▪ /analytics [дни] — расширенная аналитика\n\n"
        "Управление записями:\n"
        "▪ /edit <ID> <количество> [дата] [комментарий] — изменить запись\n"
        "▪ /delete <ID> — удалить запись\n\n"
        "Экспорт данных:\n"
        "▪ /export [дни] — выгрузить данные в Excel\n\n"
        "▪ /help — список всех команд с кратким описанием\n\n"
        "Поддержать проект:\n"
        "▪ /donate — оплатить чашку кофе ☕\n\n"
        "Используйте кнопки ниже или введите команду вручную."
    )

    await update.message.reply_text(
        commands_text,
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    )

async def add_entry(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Введите данные в формате:\n"
        "<количество яиц> [дата] [комментарий]\n\n"
        "Примеры:\n"
        "12 — добавить 12 яиц на сегодня\n"
        "12 2023-12-15 — добавить 12 яиц на 15 декабря 2023\n"
        "12 сегодня Корм поменяли — добавить 12 яиц на сегодня с комментарием"
    )


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    args = context.args
    if not args:
        # Общий список команд
        help_text = (
            "🐔 Бот для учета яйценоскости кур!\n\n"
            "Список команд:\n"
            "▪ /add — добавить запись\n"
            "▪ /stats — показать статистику\n"
            "▪ /graph — построить график\n"
            "▪ /analytics — расширенная аналитика\n"
            "▪ /edit — изменить запись\n"
            "▪ /delete — удалить запись\n"
            "▪ /export — экспорт в Excel\n"
            "▪ /help — показать справку\n\n"
            "Используйте /help <команда> для подробной информации."
        )
    else:
        # Подробное описание конкретной команды
        command = args[0].lower()
        if command == "add":
            help_text = (
                "📝 Добавление записи:\n"
                "Используйте команду /add или введите данные в формате:\n"
                "<количество яиц> [дата] [комментарий]\n\n"
                "Примеры:\n"
                "12 — добавить 12 яиц на сегодня\n"
                "12 2023-12-15 — добавить 12 яиц на 15 декабря 2023\n"
                "12 сегодня Корм поменяли — добавить 12 яиц на сегодня с комментарием"
            )
        elif command == "stats":
            help_text = (
                "📊 Статистика:\n"
                "Используйте команду /stats [дни], чтобы увидеть статистику за указанное количество дней.\n\n"
                "Примеры:\n"
                "/stats — статистика за 7 дней\n"
                "/stats 14 — статистика за 14 дней"
            )
        elif command == "graph":
            help_text = (
                "📈 График:\n"
                "Используйте команду /graph [дни], чтобы построить график яйценоскости.\n\n"
                "Примеры:\n"
                "/graph — график за 7 дней\n"
                "/graph 30 — график за 30 дней"
            )
        elif command == "analytics":
            help_text = (
                "📈 Расширенная аналитика:\n"
                "Используйте команду /analytics [дни], чтобы получить аналитику за указанное количество дней.\n\n"
                "Примеры:\n"
                "/analytics — аналитика за 7 дней\n"
                "/analytics 14 — аналитика за 14 дней"
            )
        elif command == "edit":
            help_text = (
                "✏️ Редактирование записи:\n"
                "Используйте команду /edit <ID> <количество> [дата] [комментарий], чтобы изменить запись.\n\n"
                "Примеры:\n"
                "/edit 1 15 — изменить количество яиц на 15 для записи с ID 1\n"
                "/edit 1 15 2023-12-20 Новый корм — изменить запись с ID 1"
            )
        elif command == "delete":
            help_text = (
                "❌ Удаление записи:\n"
                "Используйте команду /delete <ID>, чтобы удалить запись.\n\n"
                "Примеры:\n"
                "/delete 1 — удалить запись с ID 1"
            )
        elif command == "export":
            help_text = (
                "⬇️ Экспорт в Excel:\n"
                "Используйте команду /export [дни], чтобы получить файл со статистикой.\n\n"
                "Примеры:\n"
                "/export 5 — получить файл Excel со статистикой за 5 дней\n"
                "/export 2025-01-23 2025-02-06 — получить файл Excel со статистикой за указанный период\n"
            )

        else:
            help_text = f"❌ Команда '{command}' не найдена. Используйте /help для списка команд."

    await update.message.reply_text(help_text)

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        user_id = update.message.from_user.id
        text = update.message.text
        if text.startswith('/'):
            return

        parts = text.split(maxsplit=2)
        count = int(parts[0])  # Количество яиц

        # Обработка даты
        if len(parts) > 1 and parts[1].lower() != "сегодня":
            if not is_valid_date(parts[1]):
                await update.message.reply_text("❌ Неверный формат даты! Используйте ГГГГ-ММ-ДД.")
                return
            date = parts[1]
            notes = parts[2] if len(parts) > 2 else ""
        else:
            date = datetime.now().strftime("%Y-%m-%d")
            notes = parts[2] if len(parts) > 2 else ""

        # Добавление записи
        record_id = add_egg_record(user_id, date, count, notes)
        await update.message.reply_text(
            f"✅ Добавлено: {count} яиц\n"
            f"Дата: {date}\n"
            f"Заметка: {notes}\n"
            f"ID записи: {record_id}"
        )

    except Exception as e:
        await update.message.reply_text(
            "❌ Ошибка формата! Примеры:\n"
            "12 — добавить 12 яиц на сегодня\n"
            "12 2023-12-15 — добавить 12 яиц на 15 декабря 2023\n"
            "12 сегодня Корм поменяли — добавить 12 яиц на сегодня с комментарием"
        )
# Выгрузка в Excel
def export_to_excel(user_id, start_date=None, end_date=None):
    # Получаем данные из базы
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    query = '''SELECT date, SUM(count), GROUP_CONCAT(id)
               FROM eggs
               WHERE user_id = ?
               AND date >= ?
               AND date <= ?
               GROUP BY date
               ORDER BY date'''
    start_date = start_date or (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
    end_date = end_date or datetime.now().strftime("%Y-%m-%d")

    c.execute(query, (user_id, start_date, end_date))
    data = c.fetchall()
    conn.close()

    if not data:
        return None

    # Создаем Excel-файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Яйценоскость"
    ws.append(["Дата", "Количество яиц", "ID записей"])

    # Применяем стили к заголовкам
    for col in range(1, 4):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

   # Заполняем данные
    for date, count, ids in data:
        ws.append([date, count, ", ".join(map(str, ids))])

    # Выравниваем ID записей по правому краю
    right_alignment = Alignment(horizontal="right")
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = right_alignment

    # Автоширина столбцов
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

    # Добавляем график
    chart = LineChart()
    chart.title = "Яйценоскость"
    chart.x_axis.title = "Дата"
    chart.y_axis.title = "Количество яиц"

    data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(data) + 1)
    categories_ref = Reference(ws, min_col=1, min_row=2, max_row=len(data) + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(categories_ref)

    ws.add_chart(chart, "E2")

    # Сохраняем файл
    filename = f"egg_stats_{user_id}_{start_date}_to_{end_date}.xlsx"
    wb.save(filename)
    return filename

async def export_data(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        user_id = update.message.from_user.id
        args = context.args

        if len(args) == 2:  # Если указаны даты начала и конца
            start_date, end_date = args
            if not (is_valid_date(start_date) and is_valid_date(end_date)):
                await update.message.reply_text("❌ Неверный формат даты! Используйте ГГГГ-ММ-ДД.")
                return
        else:
            start_date = None
            end_date = None

        # Выгружаем данные в Excel
        filename = export_to_excel(user_id, start_date, end_date)
        if filename:
            await update.message.reply_document(
                document=open(filename, 'rb'),
                # caption=f"📊 Данные за период: {start_date or 'последние 7 дней'}"
                caption="Можете скачать файл с таблицей"
            )
            os.remove(filename)  # Удаляем временный файл
        else:
            await update.message.reply_text("❌ Нет данных для выгрузки.")

    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка: {str(e)}")

async def donate(update: Update, context: ContextTypes.DEFAULT_TYPE):
    donate_url = "https://pay.cloudtips.ru/p/dbed3f9a"  # Замените на вашу ссылку
    message = (
        "☕ Если вам нравится этот бот, вы можете поддержать его разработку!\n\n"
        f"👉 [Оплатить чашку кофе]({donate_url})"
    )
    await update.message.reply_text(
        message,
        parse_mode="Markdown",
        disable_web_page_preview=True
    )


if __name__ == "__main__":
    init_db()
    app = Application.builder().token(TOKEN).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("add", add_entry))
    app.add_handler(CommandHandler("edit", edit_entry))
    app.add_handler(CommandHandler("delete", delete_entry))
    app.add_handler(CommandHandler("stats", show_stats))
    app.add_handler(CommandHandler("graph", show_graph))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    app.add_handler(CommandHandler("analytics", show_analytics))
    app.add_handler(CommandHandler("export", export_data))
    app.add_handler(CommandHandler("help", help_command))
    app.add_handler(CommandHandler("donate", donate))

    app.run_polling()